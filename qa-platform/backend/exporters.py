"""Excel and Markdown exporters for the test-case registry and run results.

Everything here is built from what the store already persists — the workbook
registry and the recorded executions. Nothing is recomputed or inferred, so an
exported file and the dashboard always agree.

openpyxl is already a dependency (scripts/sync_registry.py reads the workbook
with it); write support needs no new package.
"""
from __future__ import annotations

import io
import json
import re
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------- styling
HEAD_FILL = PatternFill("solid", fgColor="1F3247")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

#: Cell tint per canonical status. Keeps a 3,000-row sheet scannable without
#: the reader having to build their own conditional formatting.
STATUS_FILL = {
    "PASS": "1E7A3C", "PASSED": "1E7A3C",
    "FAIL": "A32B22", "FAILED": "A32B22",
    "ERROR": "9A5B12", "BLOCKED": "5B2F73",
    "SKIPPED": "6B5D12", "NOT_RUN": "3A4552",
    "NOT_IMPLEMENTED": "3A4552", "MANUAL": "3A4552",
    "NOT_APPLICABLE": "3A4552", "RUNNING": "1B4C77", "QUEUED": "3A4552",
}


def _ts(value) -> str:
    if not value:
        return ""
    return datetime.fromtimestamp(
        float(value), tz=timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S")


def _flat(value) -> str:
    """Anything the store may hold in a column into a single Excel cell.

    Excel rejects control characters and caps a cell at 32,767 characters; a
    long traceback silently corrupts the file without this.
    """
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        value = ", ".join(str(v) for v in value)
    elif isinstance(value, dict):
        value = json.dumps(value, ensure_ascii=False, default=str)
    text = str(value)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", text)
    return text[:32000] + (" ...[truncated]" if len(text) > 32000 else "")


def _sheet(wb, title: str, columns: list, rows: list,
           status_cols: tuple = (), freeze: str = "A2",
           table_name: str | None = None):
    """One formatted sheet. `columns` is [(header, width), ...]."""
    ws = wb.create_sheet(title[:31])
    ws.append([c[0] for c in columns])
    for i, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = HEAD_FILL, HEAD_FONT, WRAP
    for row in rows:
        ws.append([_flat(v) for v in row])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP if columns[c - 1][1] > 40 else TOP
            if c in status_cols:
                colour = STATUS_FILL.get(str(cell.value).strip().upper())
                if colour:
                    cell.fill = PatternFill("solid", fgColor=colour)
                    cell.font = Font(color="FFFFFF", bold=True, size=10)
    ws.freeze_panes = freeze
    # An Excel table gives the reader filter dropdowns for free. Skipped on an
    # empty sheet: a table with a header row and no body row is invalid and
    # Excel refuses to open the file.
    if table_name and ws.max_row > 1:
        ref = "A1:%s%d" % (get_column_letter(len(columns)), ws.max_row)
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    return ws


def _new_workbook() -> Workbook:
    """A workbook with openpyxl's default sheet removed.

    Dropping it up front means sheets land in creation order, so Summary is
    the tab the file opens on without any move_sheet juggling.
    """
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    return wb


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------- full test-case registry
TC_COLUMNS = [
    ("Test Case ID", 18), ("Workflow", 18), ("Workflow Name", 30),
    ("Seq", 6), ("Title", 52), ("Priority", 9), ("Test Type", 16),
    ("Role", 18), ("Modules", 26), ("Suite", 16), ("Suite Name", 24),
    ("Execution Phase", 16), ("In Scope", 9),
    ("Automation Type", 16), ("Automation Status", 18), ("Automation Wave", 16),
    ("Automation Approach", 46), ("Automated By", 30), ("Related Tests", 24),
    ("Odoo 17", 12), ("Odoo 19", 12), ("Last Execution", 18),
    ("Description / User story", 60), ("Preconditions", 50), ("Steps", 70),
    ("Expected Result", 70), ("v19 Watch", 44), ("Related Features", 24),
    ("Source Notes", 34), ("Workbook", 34), ("Sheet", 22), ("Row", 7),
    ("Test Execution Row", 10),
]


def testcases_workbook(store, in_scope_only: bool = True) -> bytes:
    """Every workbook test case, one row each, plus a per-workflow summary.

    The Expected Result column is carried verbatim from the workbook, which is
    the source of truth (AUTOMATION_CONVENTIONS hard rule 2) — the exporter
    reads it and never rewrites it.
    """
    cases = store.test_cases_list(in_scope_only=in_scope_only)
    features = store.feature_summary(in_scope_only=in_scope_only)

    wb = _new_workbook()

    # --- summary sheet, so the file opens on the number people ask for first
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 58
    ws["A1"] = "DataOne 17 -> 19 — Test Case Registry"
    ws["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Exported at", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")),
        ("Scope", "In-scope workflows only" if in_scope_only
                  else "Every workflow in the workbook"),
        ("Workflows", len(features)),
        ("Test cases", len(cases)),
        ("Automatable", sum(f["automatable"] for f in features)),
        ("Automated", sum(f["automated"] for f in features)),
        ("Source workbook", cases[0]["source_workbook"] if cases else ""),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=_flat(v))

    counts: dict = {}
    for tc in cases:
        for env in ("v17", "v19"):
            counts.setdefault(env, {})
            s = tc[env + "_status"]
            counts[env][s] = counts[env].get(s, 0) + 1
    row = len(meta) + 5
    ws.cell(row=row, column=1, value="Latest status").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Status").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Odoo 17").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Odoo 19").font = Font(bold=True)
    ws.column_dimensions["C"].width = 14
    for status in sorted(set(counts.get("v17", {})) | set(counts.get("v19", {}))):
        row += 1
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=counts.get("v17", {}).get(status, 0))
        ws.cell(row=row, column=3, value=counts.get("v19", {}).get(status, 0))

    # --- one row per test case
    _sheet(wb, "Test Cases", TC_COLUMNS, [[
        tc["test_case_id"], tc["feature_id"], tc["feature_name"], tc["seq"],
        tc["title"], tc["priority"], tc["test_type"], tc["role"], tc["modules"],
        tc["suite"], tc["suite_name"], tc["execution_phase"],
        "Yes" if tc["in_scope"] else "No",
        tc["automation_type"], tc["automation_status"], tc["automation_wave"],
        tc["automation_approach"], tc["automated_test_ids"],
        tc["related_test_ids"], tc["v17_status"], tc["v19_status"],
        tc["last_execution_id"], tc["description"], tc["preconditions"],
        tc["steps"], tc["expected_result"], tc["v19_watch"],
        tc["related_features"], tc["source_notes"], tc["source_workbook"],
        tc["source_sheet"], tc["source_row"], tc["test_execution_row"],
    ] for tc in cases], status_cols=(20, 21), table_name="TestCases")

    # --- per-workflow rollup
    _sheet(wb, "Workflows", [
        ("Workflow", 18), ("Name", 40), ("Business Purpose", 70),
        ("Key Modules", 34), ("Primary Roles", 28), ("Test Cases", 11),
        ("Automatable", 12), ("Automated", 11), ("Coverage %", 11),
        ("P0", 6), ("P1", 6), ("P2", 6), ("P3", 6),
        ("v19 PASS", 10), ("v19 FAIL", 10), ("v19 BLOCKED", 12),
        ("v19 ERROR", 10), ("v17 PASS", 10), ("v17 FAIL", 10),
    ], [[
        f["feature_id"], f["name"], f["business_purpose"], f["key_modules"],
        f["primary_roles"], f["total"], f["automatable"], f["automated"],
        f["coverage_pct"],
        f["priorities"].get("P0", 0), f["priorities"].get("P1", 0),
        f["priorities"].get("P2", 0), f["priorities"].get("P3", 0),
        f["v19"].get("PASS", 0), f["v19"].get("FAIL", 0),
        f["v19"].get("BLOCKED", 0), f["v19"].get("ERROR", 0),
        f["v17"].get("PASS", 0), f["v17"].get("FAIL", 0),
    ] for f in features], table_name="Workflows")

    return _save(wb)


# ---------------------------------------------------------- run detail xlsx
def run_workbook(store, run_id: str) -> bytes:
    """One run, in full: summary, every result, every step, every assertion.

    Steps and assertions are the evidence a triage reader actually needs, so
    they get their own sheets rather than being flattened into a cell.
    """
    run = store.run(run_id)
    if not run:
        raise KeyError(run_id)
    details = [store.result(r["id"]) for r in run["results"]]
    details = [d for d in details if d]

    wb = _new_workbook()
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70
    ws["A1"] = "Run " + run_id
    ws["A1"].font = Font(bold=True, size=14)
    by_status: dict = {}
    for d in details:
        by_status[d["status"]] = by_status.get(d["status"], 0) + 1
    duration = ((run["finished_at"] or 0) - (run["started_at"] or 0)) \
        if run["started_at"] and run["finished_at"] else 0
    meta = [
        ("Label", run["label"]), ("Environment", run["env_name"]),
        ("Mode", run["mode"]), ("Status", run["status"]),
        ("Started", _ts(run["started_at"])), ("Finished", _ts(run["finished_at"])),
        ("Wall clock", "%.1f min" % (duration / 60) if duration else ""),
        ("Tests planned", run["total"]), ("Results recorded", len(details)),
        ("Exported at", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=_flat(v))
    row = len(meta) + 4
    ws.cell(row=row, column=1, value="By status").font = Font(bold=True, size=12)
    for status, n in sorted(by_status.items(), key=lambda kv: -kv[1]):
        row += 1
        cell = ws.cell(row=row, column=1, value=status)
        colour = STATUS_FILL.get(status.upper())
        if colour:
            cell.fill = PatternFill("solid", fgColor=colour)
            cell.font = Font(color="FFFFFF", bold=True)
        ws.cell(row=row, column=2, value=n)

    _sheet(wb, "Results", [
        ("Result ID", 18), ("Test ID", 24), ("Name", 54), ("Workflow", 18),
        ("Workbook TCs", 22), ("Priority", 9), ("Kind", 8), ("Status", 12),
        ("Duration (s)", 12), ("Failed Step", 46), ("Expected", 46),
        ("Actual", 46), ("Error", 70), ("Skip / Block Reason", 60),
        ("Steps", 8), ("Assertions", 11), ("Failed Assertions", 16),
        ("Started", 20), ("Finished", 20),
    ], [[
        d["id"], d["test_id"], d["name"], d["workflow"],
        ", ".join(str(x) for x in (d["traceability"] or {}).get("tc_ids", [])),
        d["priority"], d["kind"], d["status"],
        round((d["duration_ms"] or 0) / 1000, 1),
        d["failed_step"], d["expected"], d["actual"], d["error"],
        d["skip_reason"], len(d["steps"]), len(d["assertions"]),
        sum(1 for a in d["assertions"] if not a["passed"]),
        _ts(d["started_at"]), _ts(d["finished_at"]),
    ] for d in details], status_cols=(8,), table_name="Results")

    _sheet(wb, "Steps", [
        ("Test ID", 24), ("Result ID", 18), ("#", 5), ("Step", 74),
        ("Status", 12), ("Duration (s)", 12), ("Error", 80),
    ], [[
        d["test_id"], d["id"], s["idx"], s["name"], s["status"],
        round((s["duration_ms"] or 0) / 1000, 2), s["error"],
    ] for d in details for s in d["steps"]],
        status_cols=(5,), table_name="Steps")

    _sheet(wb, "Assertions", [
        ("Test ID", 24), ("Result ID", 18), ("Assertion", 60),
        ("Passed", 9), ("Expected", 60), ("Actual", 60),
    ], [[
        d["test_id"], d["id"], a["name"], "PASS" if a["passed"] else "FAIL",
        a["expected"], a["actual"],
    ] for d in details for a in d["assertions"]],
        status_cols=(4,), table_name="Assertions")

    return _save(wb)


# ------------------------------------------------------- markdown reports
def _cell(value, limit: int = 300) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")[:limit]


def _md_result(d, heading: str = "##") -> list:
    """One result, fully expanded. Shared by the per-case and per-run report."""
    tc_ids = ", ".join(str(x) for x in (d["traceability"] or {}).get("tc_ids", []))
    out = ["%s `%s` — %s" % (heading, d["test_id"], d["name"]), "",
           "| | |", "|---|---|",
           "| Status | **%s** |" % d["status"],
           "| Result ID | `%s` |" % d["id"],
           "| Workflow | %s |" % d["workflow"]]
    if tc_ids:
        out.append("| Workbook test cases | %s |" % tc_ids)
    out += ["| Priority | %s |" % d["priority"],
            "| Kind | %s |" % d["kind"],
            "| Duration | %.1fs |" % ((d["duration_ms"] or 0) / 1000),
            "| Finished | %s |" % _ts(d["finished_at"]), ""]

    if d.get("skip_reason"):
        out += [heading + "# Reason", "", d["skip_reason"], ""]
    if d.get("failed_step"):
        out += [heading + "# Failed step", "", "`%s`" % d["failed_step"], ""]
    if d.get("expected") or d.get("actual"):
        out += [heading + "# Expected vs actual", "", "```",
                "expected: %s" % d.get("expected"),
                "actual:   %s" % d.get("actual"), "```", ""]
    if d.get("error"):
        out += [heading + "# Error", "", "```", str(d["error"])[:6000], "```", ""]

    if d["steps"]:
        out += [heading + "# Steps", "",
                "| # | Step | Status | Time | Error |", "|---|---|---|---|---|"]
        for s in d["steps"]:
            mark = {"PASSED": "PASS", "FAILED": "FAIL",
                    "SKIPPED": "SKIP"}.get(s["status"], "-")
            out.append("| %s | %s | %s | %.2fs | %s |" % (
                s["idx"], _cell(s["name"] or ""), mark,
                (s["duration_ms"] or 0) / 1000, _cell(s["error"] or "", 200)))
        out.append("")

    if d["assertions"]:
        out += [heading + "# Assertions", "",
                "| | Assertion | Expected | Actual |", "|---|---|---|---|"]
        for a in d["assertions"]:
            out.append("| %s | %s | `%s` | `%s` |" % (
                "PASS" if a["passed"] else "FAIL", _cell(a["name"]),
                _cell(a["expected"]), _cell(a["actual"])))
        out.append("")

    if d["artifacts"]:
        out += [heading + "# Artifacts", ""]
        out += ["- %s: `%s`" % (a["type"], a["name"]) for a in d["artifacts"]]
        out.append("")
    return out


def result_markdown(store, result_id: str) -> str:
    d = store.result(result_id)
    if not d:
        raise KeyError(result_id)
    run = d.get("run") or {}
    head = ["# %s — %s" % (d["test_id"], d["status"]), "",
            "Run `%s` on **%s** · exported %s" % (
                d["run_id"], run.get("env_name", ""),
                datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")), ""]
    return "\n".join(head + _md_result(d, heading="##")) + "\n"


def run_markdown(store, run_id: str, only: str | None = None) -> str:
    """Every case in a run, fully expanded.

    `only` filters by status (e.g. "FAILED,ERROR") so a triage reader can pull
    just the cases that need attention without a second tool.
    """
    run = store.run(run_id)
    if not run:
        raise KeyError(run_id)
    wanted = {s.strip().upper() for s in only.split(",")} if only else None
    details = [store.result(r["id"]) for r in run["results"]]
    details = [d for d in details if d and (not wanted or d["status"] in wanted)]

    by_status: dict = {}
    for r in run["results"]:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1

    out = ["# Run %s — detailed report" % run_id, "",
           "**%s**" % run["label"], "",
           "| | |", "|---|---|",
           "| Environment | %s |" % run["env_name"],
           "| Status | %s |" % run["status"],
           "| Started | %s |" % _ts(run["started_at"]),
           "| Finished | %s |" % _ts(run["finished_at"]),
           "| Tests | %s |" % run["total"], ""]
    if by_status:
        out += ["| Status | Count |", "|---|---|"]
        out += ["| %s | %s |" % (k, v)
                for k, v in sorted(by_status.items(), key=lambda kv: -kv[1])]
        out.append("")
    if wanted:
        out += ["> Filtered to **%s** — %d of %d cases." % (
            ", ".join(sorted(wanted)), len(details), len(run["results"])), ""]

    out += ["## Contents", ""]
    out += ["- `%s` — %s" % (d["test_id"], d["status"]) for d in details]
    out.append("")
    for d in details:
        out += ["---", ""] + _md_result(d, heading="##")
    return "\n".join(out) + "\n"
