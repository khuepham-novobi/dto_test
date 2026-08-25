# -*- coding: utf-8 -*-
"""Sync the test-case registry from the DataOne Excel knowledge base.

Reads ``DataOne_v19_Test_Suite_and_Workflows_v1.0.xlsx`` (READ-ONLY — the
workbook is the source of truth and is never written) and regenerates
``data/test_registry.json``. The backend upserts that JSON into the SQLite
``test_cases`` table at startup.

Grouping model
--------------
MMG grouped test cases by *feature group* (FG-01 … FG-14). DataOne's
workbook has no feature-group sheet: the delivery unit is the **workflow**
(``DATAONE-WF-001 … WF-027``, ``Workflows`` sheet), and the ``Automation
Export`` sheet links each test case to one or more of them through
``workflow_ids``. The platform's generic ``feature_id`` slot therefore
carries the *owning workflow id*:

* a test case naming several workflows is owned by the one with the lowest
  **build order** (``Estimate and Timeline`` sheet) — it is written once, in
  the suite that is built first, and the others reference it;
* a test case naming no workflow ("— (cross-cutting)") is owned by the
  synthetic group ``DATAONE-WF-XCUT``.

Rules
-----
* ``tc_id`` is the immutable test_case_id. It is never renamed here.
* ``expected_result`` (and steps / preconditions / title / test_data /
  expected_final_state / postconditions) are copied VERBATIM from the
  workbook. Nothing in the platform may modify them — a change in the
  workbook is the only way they change.
* ``automation_type`` is derived deterministically from the workbook's
  ``automation_approach`` column (mapping below), so re-running the sync
  is idempotent.

Usage:  python scripts/sync_registry.py  [--workbook PATH]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = (Path.home() / "Downloads"
                    / "DataOne_v19_Test_Suite_and_Workflows_v1.0.xlsx")
OUT_JSON = ROOT / "data" / "test_registry.json"

SHEET = "Automation Export"
EXEC_SHEET = "Test Execution"
WORKFLOW_SHEET = "Workflows"
ESTIMATE_SHEET = "Estimate and Timeline"

# Synthetic owner for the workbook's "— (cross-cutting)" test cases: they
# name no workflow, so no workflow suite can claim them.
XCUT = "DATAONE-WF-XCUT"
XCUT_NAME = "Cross-cutting (no single workflow)"

# Scope of the current QA phase — the workflows we are generating suites for.
# Everything is imported; out-of-scope workflows are flagged so the UI can
# filter. Extend this set as later waves are picked up.
IN_SCOPE_WORKFLOWS = {
    "DATAONE-WF-020",   # Supplier Master Import from Workday
    "DATAONE-WF-003",   # Quotation revision
    "DATAONE-WF-002",   # Quotation -> sales order confirmation
    "DATAONE-WF-013",   # Customer Invoice Posting: COGS and Revenue Recognition
}

_WF_RE = re.compile(r"DATAONE-WF-\d{3}")

# Workbook automation_approach (prefix, lower-cased) -> platform
# automation_type. Order matters: the first matching prefix wins, so more
# specific prefixes are listed before the generic ones.
#
# PYTHON_UNIT          Odoo TransactionCase test inside the Odoo test runner
# ORM_INTEGRATION      odoo-bin install/upgrade + registry/log checks
# API                  connector integration test vs sandbox / mocked SFTP
# UI                   Playwright browser workflow (platform-driven)
# HTTP_CASE            Odoo HttpCase endpoint test
# TOUR                 Odoo tour (HttpCase / browser_js)
# STATIC_ANALYSIS      source-tree extraction (grep/AST) compared to the DB
# DATA_RECONCILIATION  SQL/ORM comparison v17 baseline vs v19
# PERFORMANCE          timed run against a captured Phase-0a baseline
# MANUAL               human execution required (decision gates, visual diffs)
_APPROACH_MAP = [
    ("odoo python test", "PYTHON_UNIT"),
    ("integration test against a sandbox", "API"),
    ("sql assertion script", "DATA_RECONCILIATION"),
    ("sql / orm assertion", "DATA_RECONCILIATION"),
    ("sql scan", "DATA_RECONCILIATION"),
    ("orm sweep", "PYTHON_UNIT"),
    ("odoo tour test", "TOUR"),
    ("httpcase", "HTTP_CASE"),
    ("ci boot & install", "ORM_INTEGRATION"),
    ("ci asset build", "ORM_INTEGRATION"),
    ("ci build step", "ORM_INTEGRATION"),
    ("ci:", "ORM_INTEGRATION"),
    ("static ", "STATIC_ANALYSIS"),
    ("timed run", "PERFORMANCE"),
    ("performance harness", "PERFORMANCE"),
    ("decision gate", "MANUAL"),
    ("pdf / zpl output comparison", "MANUAL"),
    ("manual ", "MANUAL"),
    ("export both reports", "MANUAL"),
    ("one-off migration task", "MANUAL"),
]

# automation_types that can never run unattended, whatever the wave says.
_HUMAN_ONLY = {"MANUAL", "PERFORMANCE"}


def _derive_automated_by() -> dict:
    """tc_id -> [platform test ids], from the live test registry.

    Read from the @test_case traceability blocks under tests/, so a suite
    that is written makes its workbook rows flip to AUTOMATED without any
    hand-maintained map.
    """
    try:
        sys.path.insert(0, str(ROOT))
        from framework import registry
        mapping: dict = {}
        for test in registry.discover():
            for raw in test.traceability.get("tc_ids", []):
                raw = str(raw)
                if "related" in raw.lower() or "gap" in raw.lower():
                    continue
                m = re.search(r"DATAONE-TC\d+", raw)
                if m:
                    mapping.setdefault(m.group(0), []).append(test.id)
        return mapping
    except Exception as exc:  # noqa: BLE001
        print(f"WARN: could not derive the automation map from tests/ "
              f"({exc}); every TC will report as not-yet-automated",
              file=sys.stderr)
        return {}


AUTOMATED_BY = _derive_automated_by()


def classify(approach: str) -> str:
    a = (approach or "").strip().lower()
    for prefix, kind in _APPROACH_MAP:
        if a.startswith(prefix):
            return kind
    return "MANUAL"          # conservative default for unknown approaches


def automation_status(tc_id: str, automation_type: str, wave: str) -> str:
    """AUTOMATED > MANUAL_ONLY > PLANNED (wave 1) > CANDIDATE (wave 2)
    > NOT_PLANNED (workbook says manual although the type is automatable)."""
    if tc_id in AUTOMATED_BY:
        return "AUTOMATED"
    if automation_type in _HUMAN_ONLY:
        return "MANUAL_ONLY"
    w = (wave or "").lower()
    if w.startswith("wave 1"):
        return "PLANNED"
    if w.startswith("wave 2"):
        return "CANDIDATE"
    if w.startswith("manual"):
        return "MANUAL_ONLY"
    return "NOT_PLANNED"


def parse_workflow_ids(raw) -> list[str]:
    """Workflow ids named by one test case, in workbook order.

    The cross-cutting marker ("— (cross-cutting)") yields an empty list.
    """
    return _WF_RE.findall(str(raw or ""))


def load_build_order(wb) -> dict:
    """workflow id -> {build_order, stage, estimate_hours, depends_on,
    timeline} from the ``Estimate and Timeline`` sheet.

    The sheet is a laid-out plan, not a table: stage headers occupy their own
    rows and the data rows start with an integer build number, so rows are
    matched by shape rather than by header lookup.
    """
    ws = wb[ESTIMATE_SHEET]
    order: dict = {}
    stage = ""
    for row in ws.iter_rows(values_only=True):
        cells = list(row) + [None] * (8 - len(row))
        first = cells[0]
        if isinstance(first, str) and first.strip().lower().startswith("stage"):
            stage = first.strip()
            continue
        if not isinstance(first, int):
            continue
        wf_ids = parse_workflow_ids(cells[1])
        if not wf_ids:
            continue
        order[wf_ids[0]] = {
            "build_order": first,
            "stage": stage,
            "area": cells[3],
            "risk": cells[4],
            "depends_on": cells[5],
            "estimate_hours": cells[6],
            "timeline": cells[7],
        }
    return order


def load_workflows(wb, build_order: dict) -> dict:
    """Workflow id -> feature-group record for the dashboard cards."""
    ws = wb[WORKFLOW_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    hdr = {h: i for i, h in enumerate(rows[0]) if h}

    def col(r, name, default=None):
        return r[hdr[name]] if name in hdr else default

    groups: dict = {}
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        wf_id = str(r[0]).strip()
        meta = build_order.get(wf_id, {})
        groups[wf_id] = {
            "feature_id": wf_id,
            "name": col(r, "Workflow"),
            "business_purpose": col(r, "Business purpose"),
            "key_modules": col(r, "Features covered"),
            "primary_roles": col(r, "Actors"),
            "in_scope": wf_id in IN_SCOPE_WORKFLOWS,
            # DataOne-specific planning metadata (unused by the generic UI,
            # consumed by scripts/gen_reports.py and the workflow page)
            "area": col(r, "Area"),
            "risk": col(r, "Risk (effective)") or col(r, "Migration risk (stated)"),
            "fail_mode": col(r, "Fail mode"),
            "primary_suites": col(r, "Primary test suites"),
            "gate_test": col(r, "Gate test — must pass before the workflow is accepted"),
            "v19_considerations": col(r, "Odoo 19 migration considerations"),
            "v17_implementation": col(r, "Odoo 17 technical implementation"),
            "build_order": meta.get("build_order"),
            "stage": meta.get("stage"),
            "estimate_hours": meta.get("estimate_hours"),
            "depends_on": meta.get("depends_on"),
            "timeline": meta.get("timeline"),
        }

    groups[XCUT] = {
        "feature_id": XCUT,
        "name": XCUT_NAME,
        "business_purpose": (
            "Test cases the workbook marks '— (cross-cutting)': they prove "
            "platform-wide behaviour (module install, security, UI labels, "
            "performance, legacy decommissioning) and name no single "
            "workflow."),
        "key_modules": None,
        "primary_roles": None,
        "in_scope": XCUT in IN_SCOPE_WORKFLOWS,
        "area": "Platform",
        "risk": None,
        "fail_mode": None,
        "primary_suites": "SMK, SEC, UIX, PERF, LEG, APR, DAT, NEW",
        "gate_test": None,
        "v19_considerations": None,
        "v17_implementation": None,
        "build_order": 0,
        "stage": "Stage 0 - Platform foundation",
        "estimate_hours": None,
        "depends_on": None,
        "timeline": None,
    }
    return groups


def owning_workflow(wf_ids: list[str], build_order: dict) -> str:
    """The workflow that owns (writes) a test case naming several.

    Lowest build order wins: the test is implemented in the suite that is
    built first, and later suites reference the same tc_id instead of
    re-implementing it. Unknown workflows sort last but keep workbook order.

    In-scope workflows are considered first. Without that preference a P0
    case naming both an in-scope and an earlier out-of-scope workflow (e.g.
    DATAONE-TC048, named by WF-013 *and* WF-008) would be owned by the
    workflow we are not building yet and silently drop out of the current
    wave — seven WF-013 P0 cases did exactly that.

    Ownership therefore tracks the *current* scope. When a later wave brings
    WF-008 in, TC048's card moves to WF-008's dashboard group, but the test
    file does not move and stays AUTOMATED: ``AUTOMATED_BY`` is derived from
    the live @test_case traceability blocks, not from this grouping.
    """
    if not wf_ids:
        return XCUT

    def rank(w):
        return (build_order.get(w, {}).get("build_order", 999), wf_ids.index(w))

    in_scope = [w for w in wf_ids if w in IN_SCOPE_WORKFLOWS]
    return min(in_scope or wf_ids, key=rank)


def load_rows(workbook: Path):
    import openpyxl
    wb = openpyxl.load_workbook(workbook, data_only=True)

    build_order = load_build_order(wb)
    groups = load_workflows(wb, build_order)

    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    hdr = {h: i for i, h in enumerate(rows[0]) if h}

    exec_ws = wb[EXEC_SHEET]
    exec_rows = list(exec_ws.iter_rows(values_only=True))
    exec_hdr = {h: i for i, h in enumerate(exec_rows[0]) if h}
    exec_row_by_tc = {
        str(r[exec_hdr["TC ID"]]).strip(): i + 2      # +2: 1-based + header
        for i, r in enumerate(exec_rows[1:])
        if r and r[exec_hdr["TC ID"]]
    }

    cases = []
    for i, r in enumerate(rows[1:]):
        if not r or not r[hdr["tc_id"]]:
            continue

        def get(col, row=r):
            return row[hdr[col]] if col in hdr else None

        tc_id = str(get("tc_id")).strip()
        wf_ids = parse_workflow_ids(get("workflow_ids"))
        owner = owning_workflow(wf_ids, build_order)
        approach = str(get("automation_approach") or "")
        # DataOne carries the wave in two columns; automation_wave is the
        # normalised one ("Wave 1 — automate now"), automation the prose one.
        wave = str(get("automation_wave") or get("automation") or "")
        auto_type = classify(approach)

        cases.append({
            "test_case_id": tc_id,
            "feature_id": owner,
            "feature_name": (groups.get(owner) or {}).get("name"),
            "in_scope": owner in IN_SCOPE_WORKFLOWS,
            "seq": get("seq"),
            "title": get("title"),                        # verbatim
            "description": get("business_purpose"),
            "feature_ref": get("feature_ids"),
            "feature": get("feature_names"),
            "feature_category": get("feature_areas"),
            "priority": get("priority"),
            "test_type": get("test_type"),
            "role": get("role_user"),
            "modules": get("modules"),
            "preconditions": get("preconditions"),        # verbatim
            "test_data": get("test_data"),                # verbatim
            "steps": get("steps"),                        # verbatim
            "expected_result": get("expected_result"),    # verbatim — source of truth
            "expected_final_state": get("expected_final_state"),   # verbatim
            "postconditions": get("postconditions"),      # verbatim
            "v19_watch": get("v19_watch"),
            "suite": get("suite"),
            "suite_name": get("suite_name"),
            "execution_phase": get("execution_phase"),
            "related_features": get("feature_names"),
            # -- DataOne workflow linkage -------------------------------
            "workflow_ids": wf_ids,
            "workflow_names": get("workflow_names"),
            "owning_workflow": owner,
            "shared_with": [w for w in wf_ids if w != owner],
            "gate_stated": get("gate_stated"),
            "hold": get("hold"),
            "phase_primary": get("phase_primary"),
            # -----------------------------------------------------------
            "automation_wave": wave,
            "automation_approach": approach,
            "automation_type": auto_type,
            "automation_status": automation_status(tc_id, auto_type, wave),
            "automated_test_ids": AUTOMATED_BY.get(tc_id, []),
            "related_test_ids": [],
            "source_notes": get("notes"),
            "source_workbook": workbook.name,
            "source_sheet": SHEET,
            "source_row": i + 2,                          # 1-based incl. header
            "test_execution_row": exec_row_by_tc.get(tc_id),
        })
    return groups, cases


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = ap.parse_args(argv)

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    groups, cases = load_rows(args.workbook)
    ids = [c["test_case_id"] for c in cases]
    if len(ids) != len(set(ids)):
        dupes = sorted({x for x in ids if ids.count(x) > 1})
        print(f"FATAL: duplicate tc_ids in workbook: {dupes}", file=sys.stderr)
        return 1

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps({
        "workbook": str(args.workbook),
        "sheet": SHEET,
        "feature_groups": list(groups.values()),
        "test_cases": cases,
    }, ensure_ascii=False, indent=1, default=str), encoding="utf-8")

    in_scope = [c for c in cases if c["in_scope"]]
    by_type: dict = {}
    for c in in_scope:
        by_type[c["automation_type"]] = by_type.get(c["automation_type"], 0) + 1
    scope_label = ", ".join(sorted(IN_SCOPE_WORKFLOWS))

    print(f"Registry written: {OUT_JSON}")
    print(f"  workflows:  {len(groups)} "
          f"({sum(1 for g in groups.values() if g['in_scope'])} in scope)")
    print(f"  test cases: {len(cases)} ({len(in_scope)} in scope)")
    print(f"  in scope:   {scope_label}")
    print(f"  in-scope automation_type: {by_type}")
    shared = [c for c in in_scope if c["shared_with"]]
    if shared:
        print(f"  shared TCs (owned here, also named by another workflow):")
        for c in shared:
            print(f"    {c['test_case_id']} -> owner {c['owning_workflow']}, "
                  f"also {', '.join(c['shared_with'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
